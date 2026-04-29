import os
import sys
import shutil
import copy
import json
import numpy as np
import pyqtgraph as pg
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from PyQt5.QtWidgets import (QApplication, QPushButton, QVBoxLayout, QWidget,
                             QLabel, QSpinBox, QHBoxLayout, QFileDialog, QInputDialog, QDialog)
from PyQt5.QtCore import Qt, pyqtSignal, QThread, QTimer, pyqtSlot
from MyMerger import CSV_merge
from pyqtgraph.parametertree import Parameter, ParameterTree
from ClassStructures.BufferProcessorClass import BufferProcessor
from ClassStructures.DeviceCommunicatorClass import DeviceCommunicator

from PyDAQmx.DAQmxConstants import (DAQmx_Val_RSE, DAQmx_Val_Volts, DAQmx_Val_Diff,
                                    DAQmx_Val_Rising, DAQmx_Val_ContSamps,
                                    DAQmx_Val_GroupByScanNumber, DAQmx_Val_Acquired_Into_Buffer,
                                    DAQmx_Val_GroupByChannel, DAQmx_Val_ChanForAllLines)

import logging
logging.basicConfig(level=logging.INFO, format='%(message)s')

def _get_daq_internal_buffer_size(task_sampling_rate):
    """
    Calculate the internal DAQ buffer size based on the task sampling rate.

    Rules:
    - 0-100 Hz: 1,000
    - 101-10,000 Hz: 10,000
    - 10,001-1,000,000 Hz: 100,000
    - >1,000,000 Hz: 1,000,000
    """
    if task_sampling_rate <= 100:
        return 1000
    elif task_sampling_rate <= 10000:
        return 10000
    elif task_sampling_rate <= 1000000:
        return 100000
    else:
        return 1000000

def _find_divisors(n):
    """Find all divisors of n, sorted"""
    divisors = set()
    for i in range(1, int(np.sqrt(n)) + 1):
        if n % i == 0:
            divisors.add(i)
            divisors.add(n // i)
    return sorted(divisors)

def _normalize_port_config(value):
    """Convert DAQmx port config constants to their symbolic names for JSON output."""
    if value == DAQmx_Val_Diff:
        return "DAQmx_Val_Diff"
    if value == DAQmx_Val_RSE:
        return "DAQmx_Val_RSE"
    return value

def _normalize_daq_tasks_for_json(DAQ_TASKS_METADATA):
    """Return a JSON-friendly deep copy of DAQ task definitions."""
    for task in DAQ_TASKS_METADATA:
        for channel in task.get("DAQ_CHANNELS", {}).values():
            if isinstance(channel, dict) and "port_config" in channel:
                channel["port_config"] = _normalize_port_config(channel["port_config"])
    return DAQ_TASKS_METADATA

def _set_group_readonly(group, readonly=True):
    group.setReadonly(readonly)
    for child in group.children():
        if child.hasChildren():
            _set_group_readonly(child, readonly)
        else:
            child.setReadonly(readonly)

# ---------------- INTERFACE AND PLOT  ----------------
class AcquisitionProgram(QWidget):

    trigger_acquisition_signal = pyqtSignal()
    start_acquisition_return_signal = pyqtSignal()
    stop_acquisition_return_signal = pyqtSignal()
    update_button_signal = pyqtSignal()
    
    def __init__(self,
                 DAQ_TASKS,
                 METADATA_COLUMNS,
                 automatic_mode=False,
                 RESISTANCE_DATA=None,
                 exp_dir=None,
                 tribu_id=None,
                 rload_id=None,
                 DAQ_CODE=(0, 0, 0, 0, 0, 1),
                 measure_time=30,
                 DAQ_USB_TRANSFER_FREQUENCY=60,  # USB transfers per second (Hz)
                 BUFFER_SAVING_TIME_INTERVAL=2.5,  # Save buffer to disk every X seconds
                 TimeWindowLength=3,  # Time window length for the plot (seconds)
                 ScreenRefreshFrequency=60,  # Screen Refresh Rate (Hz)
                 parent=None,
                 mainWindowButtons=None,
                 mainWindowParamGroups=None,
                 RelayCodeTask=None,
                 LinMotTriggerTask=None,
                 LinMotTriggerLine="Dev1/port0/line7",
                 PrepareRaspberryLine="Dev1/port0/line6",
                 RaspberryStatus_0_Line="Dev1/port1/line0",
                 RaspberryStatus_1_Line="Dev1/port1/line1",
                 RelayCodeLines="Dev1/port0/line0:5"):

        super().__init__(parent)
        self.setWindowTitle("DAQ Viewer")
        self.layout = QVBoxLayout(self)
        
        # Define the moveLinmot bool: 
        # We use a list because a bool is not referenced when passed as an argument
        self.moveLinMot = [False]
        self.xRecording = [False]
        self.LinMotTriggerTask = LinMotTriggerTask
        self.RelayCodeTask = RelayCodeTask
        
        # ---------------- CONFIG ----------------

        # DAQ acquisition parameters
        self.DAQ_USB_TRANSFER_FREQUENCY = DAQ_USB_TRANSFER_FREQUENCY
        self.BUFFER_SAVING_TIME_INTERVAL = BUFFER_SAVING_TIME_INTERVAL
        self.ACQUISITION_PARAMS = {}
        self.METADATA_COLUMNS = METADATA_COLUMNS
        self.measure_time = measure_time

        # Plot parameters
        self.TimeWindowLength = TimeWindowLength
        self.refresh_rate = int((1 / ScreenRefreshFrequency * 1000))  # Convert to milliseconds for QTimer

        # Calculate buffer sizes for each DAQ Task
        for task in DAQ_TASKS:
            # Get internal DAQ buffer size based on total sampling rate
            internal_buffer_size = _get_daq_internal_buffer_size(task["SAMPLE_RATE"])

            # Calculate initial SAMPLES_PER_CALLBACK
            initial_samples_per_callback = int(task["SAMPLE_RATE"] / DAQ_USB_TRANSFER_FREQUENCY)

            # Find divisors of internal buffer size
            divisors_of_internal_buffer = _find_divisors(internal_buffer_size)

            # Pick the divisor closest to initial_samples_per_callback
            samples_per_callback = min(divisors_of_internal_buffer,
                                      key=lambda x: abs(x - initial_samples_per_callback))

            # Log the adjustment if it changed
            if samples_per_callback != initial_samples_per_callback:
                print(f"Task {task['NAME']}: Adjusted SAMPLES_PER_CALLBACK from {initial_samples_per_callback} to {samples_per_callback}")
                print(f"  Task sampling rate: {task["SAMPLE_RATE"]} Hz, DAQ Internal buffer size: {internal_buffer_size}")

            # Calculate other parameters
            callbacks_per_buffer = int(BUFFER_SAVING_TIME_INTERVAL * DAQ_USB_TRANSFER_FREQUENCY)
            plot_buffer_size = ((task["SAMPLE_RATE"] * TimeWindowLength) // samples_per_callback) * samples_per_callback

            self.ACQUISITION_PARAMS[task["NAME"]] = {
                'SAMPLES_PER_CALLBACK': samples_per_callback,
                'BUFFER_SIZE': samples_per_callback * callbacks_per_buffer,
                'PLOT_BUFFER_SIZE': plot_buffer_size,
                'INTERNAL_BUFFER_SIZE': internal_buffer_size,
            }

        self.DAQ_TASKS = DAQ_TASKS
        self.DAQ_TASKS_METADATA = copy.deepcopy(DAQ_TASKS)
        
        self.automatic_mode = automatic_mode
        self.error_flag = False
        self.RESISTANCE_DATA = RESISTANCE_DATA
        self.iterations = 0
        self.iteration_index = 0
        self.DAQ_CODE = DAQ_CODE
        self.xClose = False
        self.mainWindowButtons = mainWindowButtons
        self.mainWindowParamGroups = mainWindowParamGroups
        self.actual_plotter = None
        self.index_pointer = None
        self.local_path = [""]

        # The order of definition is the order of saving into buffer, so we introduce an index to know the position:
        for task in DAQ_TASKS:
            for idx, (name, config) in enumerate(task["DAQ_CHANNELS"].items()):
                task["DAQ_CHANNELS"][name] = [config, idx]

        # Check RESISTANCE_DATA when automatic_mode enabled
        if self.automatic_mode:
            if self.RESISTANCE_DATA is None:
                print("Automatic mode requires RESISTANCE_DATA to be set. Exiting.")
                self.xClose = True
                return

            self.iterations = len(self.RESISTANCE_DATA) - 1

        # Request experiments directory
        if exp_dir:
            self.exp_dir = exp_dir
        else:
            print("Please select the experiment directory.")
            self.exp_dir = QFileDialog.getExistingDirectory(self, "Select Experiment Directory")
            if not self.exp_dir or not os.path.isdir(self.exp_dir):
                print("No directory selected. Exiting.")
                self.xClose = True
                return
            self.exp_dir = os.path.normpath(self.exp_dir)

        # Request TribuId
        if tribu_id:
            self.tribu_id = tribu_id
        else:
            print("Please enter TribuId.")
            self.tribu_id, ok = QInputDialog.getText(self, "Input", "Enter TribuId:")
            if not ok or not self.tribu_id:
                print("No TribuId entered. Exiting.")
                self.xClose = True
                return

        # Set rload_id if given:
        if rload_id and not self.automatic_mode:
            self.rload_id = rload_id
        else:
            self.rload_id = None

        # Define the plot widget
        self.plot_widget = pg.PlotWidget()
        self.curve = self.plot_widget.plot([], pen='y')
        self.display_data = np.array([], dtype=np.float64)

        # Acquisition control button:
        self.button = QPushButton("START LinMot")
        self.button.clicked.connect(self.trigger_acquisition)
        
        # Timer UI elements
        self.timer_label = QLabel("Duration (s):")
        self.timer_spinbox = QSpinBox()
        self.timer_spinbox.setRange(1, 86400)  # 1 sec to 24 hours
        self.timer_spinbox.setValue(measure_time)  # Default value
        self.countdown_display = QLabel("Remaining time: -")
        self.countdown_display.setAlignment(Qt.AlignRight)
        self.duration = QHBoxLayout()
        self.duration.addWidget(self.timer_label)
        self.duration.addWidget(self.timer_spinbox)
        
        # Timer setup
        self.measurement_timer = QTimer()
        self.measurement_timer.timeout.connect(self.update_countdown)
        self.remaining_seconds = 0
        self.should_save_data = False
        
        # Buffer processor and thread for each DAQ Task
        self.buffer_processors = []
        self.thread_savers = []
        self.task_names = []
        for n, task in enumerate(self.DAQ_TASKS):
            self.buffer_processors.append(BufferProcessor(fs=task["SAMPLE_RATE"],
                                                          mainWindowReference=self,
                                                          channel_config=task["DAQ_CHANNELS"],
                                                          task_name=task["NAME"],
                                                          task_type=task["TYPE"]))
            self.thread_savers.append(QThread())
            self.task_names.append(task["NAME"])
            self.buffer_processors[n].moveToThread(self.thread_savers[n])
            self.thread_savers[n].start()

        # Raspberry Communicator + DAQ Analog Task (2 threads):
        self.dev_comunicator = DeviceCommunicator(mainWindowReference=self,
                                                  RelayCodeTask=RelayCodeTask,
                                                  LinMotTriggerTask=LinMotTriggerTask,
                                                  LinMotTriggerLine=LinMotTriggerLine,
                                                  PrepareRaspberryLine=PrepareRaspberryLine,
                                                  RaspberryStatus_0_Line=RaspberryStatus_0_Line,
                                                  RaspberryStatus_1_Line=RaspberryStatus_1_Line,
                                                  RelayCodeLines=RelayCodeLines)

        self.thread_communicator = QThread()
        self.dev_comunicator.moveToThread(self.thread_communicator)
        self.thread_communicator.start()
        
        # Plot update timer
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_plot)
        self.timer.start(self.refresh_rate)

        # Signal management
        self.trigger_acquisition_signal.connect(self.trigger_acquisition)
        self.start_acquisition_return_signal.connect(self.start_acquisition_return)
        self.stop_acquisition_return_signal.connect(self.stop_acquisition_return)
        self.update_button_signal.connect(self.update_button)

        # Insert the DAQ Task reference in each channel and generate a dictionary with all channels
        total_tasks = {}
        for n, task in enumerate(self.DAQ_TASKS):
            for channel in task["DAQ_CHANNELS"].keys():
                task["DAQ_CHANNELS"][channel].insert(-1, self.dev_comunicator.AcquisitionTasks[n])
            total_tasks = total_tasks | task["DAQ_CHANNELS"]

        # Signal Selector
        self.signal_selector = Parameter.create(name='Signal', type='list', limits=total_tasks)
        self.signal_selector.sigValueChanged.connect(self._update_DAQ_Plot_Buffer)

        # Now we can assign the signal selector to the DAQs:
        for task in self.dev_comunicator.AcquisitionTasks:
            task.data_column_selector = self.signal_selector

        # Set the layout
        self.tree = ParameterTree()
        self.tree.setParameters(self.signal_selector, showTop=True)

        # Button to open experiment defaults editor (in separate dialog)
        self.edit_defaults_button = QPushButton("Edit Experiment Parameters")
        self.edit_defaults_button.clicked.connect(self.open_defaults_dialog)

        # Build the hidden Parameter group (not shown until dialog opens)
        children = []
        for col, meta in self.METADATA_COLUMNS.items():
            if col in ('Date', 'ReadingTime (s)'):
                continue  # Auto-generated and must not be editable in the dialog.
            ptype = meta.get('type', 'str')
            if ptype not in ('str', 'int', 'float', 'bool', 'list'):
                ptype = 'str'
            children.append({'name': col, 'type': ptype, 'value': meta.get('default', '')})

        self.metadata_param_tree = Parameter.create(name='ExperimentDefaults', type='group', children=children)

        # Add widgets to the layout (sorted from top to bottom)
        self.layout.addWidget(self.tree)
        self.layout.addWidget(self.edit_defaults_button)
        self.layout.addWidget(self.button)
        self.layout.addWidget(self.plot_widget)
        self.layout.addLayout(self.duration)
        self.layout.addWidget(self.countdown_display)

        if self.automatic_mode:
            print("Starting acquisition in automatic mode.")
            self.trigger_acquisition()

    def _update_DAQ_Plot_Buffer(self, param, value):
        # Obtain the DAQ Task Reference
        DAQ_Task_Reference = value[1]

        # Disconnect the plotter from the screen
        self.actual_plotter = None

        # Reshape the display_data buffer
        if self.display_data.size != DAQ_Task_Reference.PLOT_BUFFER_SIZE:
            self.display_data = np.empty(DAQ_Task_Reference.PLOT_BUFFER_SIZE, dtype=np.float64)

        # Calculate the array index corresponding to the selected signal
        self.index_pointer = self.signal_selector.value()[-1]

        # Select the actual plotter
        self.actual_plotter = DAQ_Task_Reference

    def flush_screen(self):
        self.actual_plotter = None
        self.curve.setData([])

    def update_countdown(self):

        if self.dev_comunicator.is_rb_connected:
            # It takes aprox 10e-5 seconds to read the status bits
            status_bit_0 = self.dev_comunicator.DI_task_Raspberry_status_0.read_line()
            status_bit_1 = self.dev_comunicator.DI_task_Raspberry_status_1.read_line()

            if not (status_bit_0 == 1 and status_bit_1 == 0):
                if not self.error_flag:
                    # An error has occurred during the data acquisition
                    print(f"\033[91mError, during the acquisition, the Raspberry sent an error code:\033[0m", f"{status_bit_0}{status_bit_1}")
                    self.error_flag = True
                    self.trigger_acquisition()
            
        if not self.error_flag:
            if self.remaining_seconds > 0 and self.moveLinMot[0]:
                self.remaining_seconds -= 1
                self.countdown_display.setText(f"Remaining time: {self.remaining_seconds} s")
            else:
                self.measurement_timer.stop()
                self.countdown_display.setText("Remaining time: -")
                self.should_save_data = True
                self.trigger_acquisition()
    
    def update_plot(self):
        if self.actual_plotter is None:
            return
        if self.display_data.size == 0:
            return

        plot_buffer = self.actual_plotter.plot_buffer
        plot_buffer_size = self.actual_plotter.PLOT_BUFFER_SIZE
        write_index = self.actual_plotter.write_index

        tail_size = plot_buffer_size - write_index
        self.display_data[:tail_size] = plot_buffer[write_index:, self.index_pointer]
        self.display_data[tail_size:] = plot_buffer[:write_index, self.index_pointer]

        self.curve.setData(self.display_data)


    @pyqtSlot()
    def start_acquisition_return(self):

        if not self.error_flag:
            self.remaining_seconds = self.timer_spinbox.value()
            self.measure_time = self.timer_spinbox.value()
            self.countdown_display.setText(f"Remaining time: {self.remaining_seconds} s")
            self.measurement_timer.start(1000)  # 1 sec
            self.should_save_data = False

            # Assign the plotter to the selected signal
            self.actual_plotter = self.signal_selector.value()[-2]

            self.update_button()
            print("The acquisition has started successfully!")

        else:
            # Close the opened files
            for processor in self.buffer_processors:
                processor.close_file()

            # Delete the folder created
            shutil.rmtree(self.local_path[0])

    @pyqtSlot()
    def stop_acquisition_return(self):

        # Clean the screen
        self.flush_screen()

        # Close the opened files
        for processor in self.buffer_processors:
            processor.close_file()

        if not self.error_flag:
            if self.should_save_data:

                # Convert the DAQ binary files to pandas dataframes
                for processor in self.buffer_processors:
                    processor.Binary_to_Pickle()

                # Merge the LinMot CSV files
                if self.dev_comunicator.is_rb_connected:
                    self.motor_file = CSV_merge(folder_path=self.local_path[0], exp_id=self.exp_id)
                else:
                    self.motor_file = ""

                # Save the metadata
                experiment_metadata = self._build_experiment_metadata()
                error_code = self.save_metadata(experiment_metadata)

                if error_code == 0:
                    print("Experiment ended succesfully!")
            else:
                print("Experiment interrupted.")

            if self.automatic_mode:
                if self.iteration_index <= self.iterations:
                    self.trigger_acquisition()
                else:
                    print("All iterations finished, exiting.")
                    self.close()
            else:
                self.update_button()
        else:
            self.update_button()
            print("Experiment interrupted due to an error.")

        # Delete temporal files if no error, if error, delete all
        if not self.error_flag:
            for processor in self.buffer_processors:
                processor.remove_file()
        else:
            # Delete all files
            shutil.rmtree(self.local_path[0])

    def update_button(self):
        self.button.setText("STOP LinMot" if self.moveLinMot[0] else "START LinMot")

    @pyqtSlot()
    def trigger_acquisition(self):

        if self.sender() == self.button and self.automatic_mode:
            print("Automatic mode has been disabled, stopping acquisition.")
            self.automatic_mode = False

        if self.error_flag:
            if self.sender() == self.button:
                self.error_flag = False
            else:
                # STOP ADQUISITION
                self.measurement_timer.stop()
                self.countdown_display.setText("Remaining time: -")
                self.dev_comunicator.stop_acquisition_signal.emit()
                print("Stopped acquisition due to an error.")
                return

        if self.automatic_mode:
            print(f"Starting the iteration {self.iteration_index} of {self.iterations}")

        if self.moveLinMot[0]:
            # STOP ADQUISITION
            self.measurement_timer.stop()
            self.countdown_display.setText("Remaining time: -")
            self.dev_comunicator.stop_acquisition_signal.emit()
        else:
            # START ADQUISITION
            if self.automatic_mode:
                self.rload_id = self.RESISTANCE_DATA[self.iteration_index]["RLOAD_ID"]
            elif not self.rload_id:
                print("\nPlease enter RloadId")
                self.rload_id, ok = QInputDialog.getText(self, "Input", "Enter RloadId:")
                if not ok or not self.rload_id:
                    print("No RloadId entered. Operation canceled.")
                    return

            self.date_now = datetime.now().strftime("%d%m%Y_%H%M%S")
            self.exp_id = f"{self.date_now}-{self.tribu_id}-{self.rload_id}"

            # Create necessary folders and define the saving path
            os.makedirs(os.path.join(self.exp_dir, "RawData"), exist_ok=True)
            self.local_path[0] = os.path.join(self.exp_dir, "RawData", self.exp_id)
            os.makedirs(self.local_path[0], exist_ok=True)

            # Open the files to save the data
            for processor in self.buffer_processors:
                processor.open_file()

            self.dev_comunicator.start_acquisition_signal.emit()

    def open_defaults_dialog(self):
        """Open a modal dialog to edit experiment default values (save or cancel)."""
        if not hasattr(self, 'metadata_param_tree') or self.metadata_param_tree is None:
            print("Experiment defaults schema unavailable.")
            return

        # Backup current values
        orig = {}
        for col in list(self.METADATA_COLUMNS.keys()):
            if col in ('Date', 'ReadingTime (s)'):
                continue
            try:
                p = self.metadata_param_tree.param(col)
                orig[col] = p.value() if p is not None else None
            except Exception:
                orig[col] = None

        dlg = QDialog(self)
        dlg.setWindowTitle("Edit Experiment Defaults")
        dlg_layout = QVBoxLayout(dlg)

        tree = ParameterTree()
        tree.setParameters(self.metadata_param_tree, showTop=True)
        dlg_layout.addWidget(tree)

        btn_layout = QHBoxLayout()
        btn_save = QPushButton("Save")
        btn_cancel = QPushButton("Cancel")
        btn_layout.addWidget(btn_save)
        btn_layout.addWidget(btn_cancel)
        dlg_layout.addLayout(btn_layout)

        def on_cancel():
            # restore originals
            for col, val in orig.items():
                try:
                    p = self.metadata_param_tree.param(col)
                    if p is not None:
                        p.setValue(val)
                except Exception:
                    pass
            dlg.reject()

        def on_save():
            dlg.accept()

        btn_cancel.clicked.connect(on_cancel)
        btn_save.clicked.connect(on_save)

        dlg.exec_()

    def _build_experiment_metadata(self):
        """Build the experiment metadata split into Excel and JSON payloads."""
        try:
            date_value = datetime.strptime(self.date_now, "%d%m%Y_%H%M%S").date()
        except Exception:
            date_value = self.date_now

        excel_metadata = {}

        # Generate the Excel metadata dictionary
        for col in self.METADATA_COLUMNS.keys():
            if col == 'Date':
                excel_metadata[col] = date_value  # always auto-generated, must be the third column
                continue

            if col == 'ReadingTime (s)':
                excel_metadata[col] = self.measure_time  # determined by software
                continue

            param = self.metadata_param_tree.param(col)
            if param is not None:
                val = param.value()
            else:
                val = None

            excel_metadata[col] = val

        json_metadata = {
            "ExperimentId": self.exp_id,
            "RaspberryConnected": self.dev_comunicator.is_rb_connected,
            "DAQTasks": _normalize_daq_tasks_for_json(self.DAQ_TASKS_METADATA),
        }

        return {
            "excel_metadata": excel_metadata,
            "json_metadata": json_metadata,
        }

    def _normalize_cell_value(self, value):
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value)
        return value

    def _set_column_widths(self, ws):
        """Set column widths and style headers for better readability in Excel."""
        from openpyxl.utils import get_column_letter
        # Define header styling: light blue background with white text
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_idx, header in enumerate(list(self.METADATA_COLUMNS.keys()), start=1):
            col_letter = get_column_letter(col_idx)
            # Set width based on header length, with a minimum of 15 and maximum of 50
            width = max(15, min(len(header) + 3, 50))
            ws.column_dimensions[col_letter].width = width
            
            # Apply header styling to the first row
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = header_alignment

        # Set a fixed row height for consistency
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 15

    def _apply_center_alignment(self, ws):
        """Apply center alignment to all data cells in the worksheet (excluding header row)."""
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Start from row 2 to skip the header row
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = center_alignment

    def _ensure_experiment_workbook(self, file_path):
        """Create or normalize the workbook so it contains exactly the metadata columns."""
        if not os.path.isfile(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "MetadataSheet"
            for col_idx, header in enumerate(list(self.METADATA_COLUMNS.keys()), start=1):
                ws.cell(row=1, column=col_idx, value=header)
            # Set column widths and apply center alignment
            self._set_column_widths(ws)
            self._apply_center_alignment(ws)
            wb.save(file_path)
            return

        wb = load_workbook(file_path)
        ws = wb.active
        existing_headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        existing_headers = [str(h).strip() for h in existing_headers if h not in (None, "")]

        if existing_headers == list(self.METADATA_COLUMNS.keys()):
            return

        # Rebuild the workbook to enforce the exact ODS schema.
        rows = []
        header_map = {header: idx + 1 for idx, header in enumerate(existing_headers)}
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for header in existing_headers:
                col_idx = header_map[header]
                row_data[header] = ws.cell(row=row_idx, column=col_idx).value
            rows.append(row_data)

        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = "MetadataSheet"
        for col_idx, header in enumerate(list(self.METADATA_COLUMNS.keys()), start=1):
            new_ws.cell(row=1, column=col_idx, value=header)
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, header in enumerate(self.METADATA_COLUMNS, start=1):
                new_ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
        # Set column widths and apply center alignment
        self._set_column_widths(new_ws)
        self._apply_center_alignment(new_ws)
        new_wb.save(file_path)

    def save_metadata(self, experiment_data, base_filename="Experiments"):
        """Create/update Experiments.xlsx schema and append one row from input dictionary, and creates a JSON File"""
        if not isinstance(experiment_data, dict) or not experiment_data:
            print(f"\033[91mCould not save experiment row: experiment_data must be a non-empty dictionary. \033[0m")
            return 1

        excel_metadata = experiment_data.get("excel_metadata")
        json_metadata = experiment_data.get("json_metadata")

        if not isinstance(excel_metadata, dict) or not isinstance(json_metadata, dict):
            print(f"\033[91mCould not save experiment row: experiment_data must contain 'excel_metadata' and 'json_metadata' dictionaries. \033[0m")
            return 1

        folder_path = os.path.dirname(self.local_path[0])
        if not folder_path:
            print(f"\033[91mCould not save experiment row: invalid experiment folder path. \033[0m")
            return 1

        os.makedirs(folder_path, exist_ok=True)
        file_path = os.path.join(folder_path, f"{base_filename}.xlsx")
        json_path = os.path.join(self.local_path[0], "experiment_metadata.json")

        error_code = 0

        try:
            # Save Excel row
            self._ensure_experiment_workbook(file_path)
            wb = load_workbook(file_path)
            ws = wb.active

            header_to_col = {header: idx + 1 for idx, header in enumerate(self.METADATA_COLUMNS)}
            write_row = ws.max_row + 1

            sheet_row = {}
            for header in list(self.METADATA_COLUMNS.keys()):
                default = self.METADATA_COLUMNS.get(header, {}).get('default', "")
                val = excel_metadata.get(header, default)
                sheet_row[header] = self._normalize_cell_value(val)

            for header, value in sheet_row.items():
                ws.cell(row=write_row, column=header_to_col[header], value=value)

            # Set column widths and apply center alignment
            self._set_column_widths(ws)
            self._apply_center_alignment(ws)

            # Save Excel File
            wb.save(file_path)

        except Exception as e:
            print(f"\033[91mCould not save experiment row in {file_path}: {e} \033[0m")
            error_code = 1

        try:
            # Save JSON File
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(json_metadata, f, ensure_ascii=False, indent=2, default=str)
        except Exception as e:
            print(f"\033[91mCould not JSON File in {json_path}: {e} \033[0m")
            error_code = 1

        return error_code

    def closeEvent(self, event):

        print("\nClosing DAQ Viewer")

        if not self.xClose:
            # Cleanup DAQ tasks
            for task in self.dev_comunicator.AcquisitionTasks:
                task.StopTask()
                task.ClearTask()

            self.dev_comunicator.DO_task_LinMotTrigger.set_line(0)
            if not self.LinMotTriggerTask:
                self.dev_comunicator.DO_task_LinMotTrigger.StopTask()
                self.dev_comunicator.DO_task_LinMotTrigger.ClearTask()

            self.dev_comunicator.DO_task_RelayCode.set_lines([0,0,0,0,0,0])
            if not self.RelayCodeTask:
                self.dev_comunicator.DO_task_RelayCode.StopTask()
                self.dev_comunicator.DO_task_RelayCode.ClearTask()

            self.dev_comunicator.DO_task_PrepareRaspberry.set_line(0)
            self.dev_comunicator.DO_task_PrepareRaspberry.StopTask()
            self.dev_comunicator.DO_task_PrepareRaspberry.ClearTask()

            self.dev_comunicator.DI_task_Raspberry_status_0.StopTask()
            self.dev_comunicator.DI_task_Raspberry_status_0.ClearTask()

            self.dev_comunicator.DI_task_Raspberry_status_1.StopTask()
            self.dev_comunicator.DI_task_Raspberry_status_1.ClearTask()

            for thread in self.thread_savers:
                thread.quit()
                thread.wait()

            self.thread_communicator.quit()
            self.thread_communicator.wait()

            if self.moveLinMot[0] and os.path.isdir(self.local_path[0]):
                shutil.rmtree(self.local_path[0])
                print(f"Temporary folder {self.local_path[0]} deleted on exit.")

        if self.mainWindowButtons:
            for button in self.mainWindowButtons.values():
                button.setEnabled(True)

        if self.mainWindowParamGroups:
            for param_group in self.mainWindowParamGroups.values():
                _set_group_readonly(param_group, readonly=False)


        # Use the accept() method of the class QCloseEvent to close the QWidget (if not desired use the ignore() method)
        event.accept()

    def showEvent(self, event):
        super().showEvent(event)
        if self.xClose:
            # Close the QWidget as soon as it is created adding in the event loop a QTimer event
            QTimer.singleShot(0, self.close)

# ---------------- MAIN ----------------
if __name__ == '__main__':

    DAQ_TASKS = [
        {
            "NAME": "Dev1",
            "SAMPLE_RATE": 10000,

            "DAQ_CHANNELS":{
                "Voltage": {"port": "Dev1/ai3", "port_config": DAQmx_Val_Diff, "conversion_factor": None},
                # "Current": {"port": "Dev1/ai2", "port_config": DAQmx_Val_RSE, "conversion_factor": 1},
            },

            # "TRIGGER_SOURCE": "PFI0",
            "TRIGGER_SOURCE": None,

            "TYPE": "analog"
        },

        {
            "NAME": "Dev2",
            "SAMPLE_RATE": 10000,

            "DAQ_CHANNELS":{
                # "LinMot_Enable": {"port":"Dev2/ai0", "port_config":DAQmx_Val_RSE, "conversion_factor": None},
                # "LinMot_Up_Down": {"port":"Dev2/ai1", "port_config":DAQmx_Val_RSE, "conversion_factor": None}
                "LinMot_Enable": {"port":"Dev2/port0/line0", "port_config":None, "conversion_factor": None},
                "LinMot_Up_Down": {"port":"Dev2/port0/line1", "port_config":None, "conversion_factor": None}
            },

            # "TRIGGER_SOURCE": "PFI0",
            "TRIGGER_SOURCE": None,
            "TYPE": "digital"
        }
    ]

    METADATA_COLUMNS = {
        "TribuId": {"default": "Nylon6-PDMS", "type": "str", "limits": None},
        "Keithley Used": {"default": False, "type": "bool", "limits": None},
        "Date": {"default": None, "type": "date", "limits": None},
        "Capacitorld": {"default": "none", "type": "str", "limits": None},
        "RloadId": {"default": "R100", "type": "str", "limits": None},
        "InitialPosition": {"default": -54.0, "type": "float", "limits": None},
        "FinalPosition": {"default": -59.4, "type": "float", "limits": None},
        "MeasuredParameterMotor": {"default": "Pos(mm), F(N), Target Force(N)", "type": "str", "limits": None},
        "ReadingTime (s)": {"default": 0, "type": "int", "limits": [0, None]},
        "Electrode rGO": {"default": False, "type": "bool", "limits": None},
        "MotorSpeedDown(m/s)": {"default": 0.5, "type": "float", "limits": None},
        "MotorAccelerationDown(m/s)": {"default": 0.5, "type": "float", "limits": None},
        "MotorDecelerationDown(m/s)": {"default": 0.5, "type": "float", "limits": None},
        "Motor Speed Up(m/s)": {"default": 0.5, "type": "float", "limits": None},
        "Motor AccelerationUp(m/s)": {"default": 0.5, "type": "float", "limits": None},
        "MotorDecelerationUp(m/s)": {"default": 0.5, "type": "float", "limits": None},
        "MotorForceMax": {"default": 0.0, "type": "float", "limits": None},
        "Notes": {"default": "", "type": "str", "limits": None},
    }

    resistance_list = [{'VALUE': 5,
                      'ATENUATION': 1.0,
                      'DAQ_CODE': [0, 0, 0, 0, 0, 1],
                      'RLOAD_ID': 'Resistance 0'},
                     {'VALUE': 25,
                      'ATENUATION': 1.8,
                      'DAQ_CODE': [0, 0, 1, 0, 0, 1],
                      'RLOAD_ID': 'Resistance 4'},
                     {'VALUE': 35,
                      'ATENUATION': 2.2,
                      'DAQ_CODE': [0, 1, 1, 0, 0, 1],
                      'RLOAD_ID': 'Resistance 6'}]

    app = QApplication(sys.argv)

    # To debug, use this
    debug = True
    tribo_lab = True
    if debug:
        if tribo_lab:
            exp_dir = r"C:\Users\mmartic\Desktop\RogerTest"
        else:
            exp_dir = r"C:\Users\rpieres\Desktop\Test"
        tribu_id = "PDMSvsNylon"
        rload_id = "10"
    else:
        exp_dir = None
        tribu_id = None
        rload_id = None

    window = AcquisitionProgram(DAQ_TASKS=DAQ_TASKS,
                                METADATA_COLUMNS=METADATA_COLUMNS,
                                automatic_mode=False,
                                RESISTANCE_DATA=resistance_list,
                                measure_time=1,
                                exp_dir=exp_dir,
                                tribu_id=tribu_id,
                                rload_id=rload_id)
    window.show()
    sys.exit(app.exec_())
