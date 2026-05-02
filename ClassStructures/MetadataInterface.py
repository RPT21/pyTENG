from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from datetime import datetime
import json
import os

from PyDAQmx.DAQmxConstants import (DAQmx_Val_RSE, DAQmx_Val_Volts, DAQmx_Val_Diff,
                                    DAQmx_Val_Rising, DAQmx_Val_ContSamps,
                                    DAQmx_Val_GroupByScanNumber, DAQmx_Val_Acquired_Into_Buffer,
                                    DAQmx_Val_GroupByChannel, DAQmx_Val_ChanForAllLines)

class MetadataInterface:

    def __init__(self, mainWindowReference):
        self.mainWindow = mainWindowReference

    def _normalize_port_config(self, value):
        """Convert DAQmx port config constants to their symbolic names for JSON output."""
        if value == DAQmx_Val_Diff:
            return "DAQmx_Val_Diff"
        if value == DAQmx_Val_RSE:
            return "DAQmx_Val_RSE"
        return value

    def _normalize_daq_tasks_for_json(self, DAQ_TASKS_METADATA):
        """Return a JSON-friendly deep copy of DAQ task definitions."""
        for task in DAQ_TASKS_METADATA:
            for channel in task.get("DAQ_CHANNELS", {}).values():
                if isinstance(channel, dict) and "port_config" in channel:
                    channel["port_config"] = self._normalize_port_config(channel["port_config"])
        return DAQ_TASKS_METADATA

    def build_experiment_metadata(self):
        """Build the experiment metadata split into Excel and JSON payloads."""
        try:
            date_value = datetime.strptime(self.mainWindow.date_now, "%d%m%Y_%H%M%S").date()
        except Exception:
            date_value = self.mainWindow.date_now

        excel_metadata = {}

        # Generate the Excel metadata dictionary
        for col in self.mainWindow.METADATA_COLUMNS.keys():
            if col == 'Date':
                excel_metadata[col] = date_value  # always auto-generated, must be the third column
                continue

            if col == 'ReadingTime (s)':
                excel_metadata[col] = self.mainWindow.measure_time  # determined by software
                continue

            param = self.mainWindow.ExpConfigWindow.metadata_param_tree.param(col)
            if param is not None:
                val = param.value()
            else:
                val = None

            excel_metadata[col] = val

        json_metadata = {
            "ExperimentId": self.mainWindow.exp_id,
            "DAQProfile": self.mainWindow.active_daq_profile_name,
            "RaspberryConnected": self.mainWindow.dev_comunicator.is_rb_connected,
            "DAQTasks": self._normalize_daq_tasks_for_json(self.mainWindow.DAQ_TASKS_METADATA),
        }

        return {
            "excel_metadata": excel_metadata,
            "json_metadata": json_metadata,
        }

    def _normalize_cell_value(self, value):
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value)
        return value

    def _set_column_widths(self, ws):
        """Set column widths and style headers for better readability in Excel."""
        from openpyxl.utils import get_column_letter
        # Define header styling: light blue background with white text
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx, header in enumerate(list(self.mainWindow.METADATA_COLUMNS.keys()), start=1):
            col_letter = get_column_letter(col_idx)
            # Set width based on header length, with a minimum of 15 and maximum of 50
            width = max(15, min(len(header) + 3, 50))
            ws.column_dimensions[col_letter].width = width

            # Apply header styling to the first row
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = header_alignment

        # Set a fixed row height for consistency
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 15

    def _apply_center_alignment(self, ws):
        """Apply center alignment to all data cells in the worksheet (excluding header row)."""
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Start from row 2 to skip the header row
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = center_alignment

    def _ensure_experiment_workbook(self, file_path):
        """Create or normalize the workbook so it contains exactly the metadata columns."""
        if not os.path.isfile(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "MetadataSheet"
            for col_idx, header in enumerate(list(self.mainWindow.METADATA_COLUMNS.keys()), start=1):
                ws.cell(row=1, column=col_idx, value=header)
            # Set column widths and apply center alignment
            self._set_column_widths(ws)
            self._apply_center_alignment(ws)
            wb.save(file_path)
            return

        wb = load_workbook(file_path)
        ws = wb.active
        existing_headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        existing_headers = [str(h).strip() for h in existing_headers if h not in (None, "")]

        if existing_headers == list(self.mainWindow.METADATA_COLUMNS.keys()):
            return

        # Rebuild the workbook to enforce the exact ODS schema.
        rows = []
        header_map = {header: idx + 1 for idx, header in enumerate(existing_headers)}
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for header in existing_headers:
                col_idx = header_map[header]
                row_data[header] = ws.cell(row=row_idx, column=col_idx).value
            rows.append(row_data)

        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = "MetadataSheet"
        for col_idx, header in enumerate(list(self.mainWindow.METADATA_COLUMNS.keys()), start=1):
            new_ws.cell(row=1, column=col_idx, value=header)
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, header in enumerate(self.mainWindow.METADATA_COLUMNS, start=1):
                new_ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
        # Set column widths and apply center alignment
        self._set_column_widths(new_ws)
        self._apply_center_alignment(new_ws)
        new_wb.save(file_path)

    def save_metadata(self, experiment_data, base_filename="Experiments"):
        """Create/update Experiments.xlsx schema and append one row from input dictionary, and creates a JSON File"""
        if not isinstance(experiment_data, dict) or not experiment_data:
            print(f"\033[91mCould not save experiment row: experiment_data must be a non-empty dictionary. \033[0m")
            return 1

        excel_metadata = experiment_data.get("excel_metadata")
        json_metadata = experiment_data.get("json_metadata")

        if not isinstance(excel_metadata, dict) or not isinstance(json_metadata, dict):
            print(
                f"\033[91mCould not save experiment row: experiment_data must contain 'excel_metadata' and 'json_metadata' dictionaries. \033[0m")
            return 1

        folder_path = os.path.dirname(self.mainWindow.local_path[0])
        if not folder_path:
            print(f"\033[91mCould not save experiment row: invalid experiment folder path. \033[0m")
            return 1

        os.makedirs(folder_path, exist_ok=True)
        file_path = os.path.join(folder_path, f"{base_filename}.xlsx")
        json_path = os.path.join(self.mainWindow.local_path[0], "experiment_metadata.json")

        error_code = 0

        try:
            # Save Excel row
            self._ensure_experiment_workbook(file_path)
            wb = load_workbook(file_path)
            ws = wb.active

            header_to_col = {header: idx + 1 for idx, header in enumerate(self.mainWindow.METADATA_COLUMNS)}
            write_row = ws.max_row + 1

            sheet_row = {}
            for header in list(self.mainWindow.METADATA_COLUMNS.keys()):
                default = self.mainWindow.METADATA_COLUMNS.get(header, {}).get('default', "")
                val = excel_metadata.get(header, default)
                sheet_row[header] = self._normalize_cell_value(val)

            for header, value in sheet_row.items():
                ws.cell(row=write_row, column=header_to_col[header], value=value)

            # Set column widths and apply center alignment
            self._set_column_widths(ws)
            self._apply_center_alignment(ws)

            # Save Excel File
            wb.save(file_path)

        except Exception as e:
            print(f"\033[91mCould not save experiment row in {file_path}: {e} \033[0m")
            error_code = 1

        try:
            # Save JSON File
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(json_metadata, f, ensure_ascii=False, indent=2, default=str)
        except Exception as e:
            print(f"\033[91mCould not JSON File in {json_path}: {e} \033[0m")
            error_code = 1

        return error_code